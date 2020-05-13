import geocoder #pip install
import openpyxl #pip install
import time

filename = 'hotel2.xlsx'
exelFile = openpyxl.load_workbook(filename)
ws = exelFile.active
ws.delete_cols(2)
ws.delete_cols(3)
ws.delete_cols(4)
ws.delete_cols(5)

sheet = exelFile.worksheets[0]

rowCount=1

for row in sheet.rows:
    g = geocoder.google(row[0].value)

    #if g.status == "OVER_QUERY_LIMIT":
	#	while True:
	#		print("Fail ===> Try Again")
	#		time.sleep(0.5)
	#		g = geocoder.google(row[0].value)
	#		if g.status == "OK":
	#			break
    lat_cell = sheet.cell(row = rowCount, column = 2)
    lng_cell = sheet.cell(row = rowCount, column = 3)
    
    geo = g.latlng
    
    print(geo)
    lat_cell.value = geo[0]
    lng_cell.value = geo[1]
    rowCount = rowCount + 1

exelFile.save("address.xlsx")
